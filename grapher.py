
import tkinter as tk
from tkinter import filedialog
from openpyxl import *

print("Starting up...")
earth_temp = input('Input the average measured at the surface of the earth... ')
space_temp = input('Input the average measured in empty space... ')

earth_temp = float(earth_temp)
space_temp = float(space_temp)

assert earth_temp > space_temp

slope_correction = (300-3)/(earth_temp - space_temp)
intercept = 300 - (slope_correction * earth_temp)

print("\nThe data correction line is y = "+str(slope_correction)+"*x + "+str(intercept))
print("\nPlease select the .rad file with the data you want to graph")

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()

print('\nfile selected: '+file_path)

rad_file = open(file_path, 'r')
lines = rad_file.readlines()
rad_file.close()
line_data = lines[1].split(' ')

num_samples = len(lines) 

print('\nbeginning data separation...')

for sample_num in range(1, num_samples):
	lines[sample_num] = lines[sample_num].split(' ')
	num_spaces = lines[sample_num].count('')
	
	for i in range(num_spaces):
		lines[sample_num].remove('')


#debug lines
#print('\n', lines[1])

starting_freq = float(lines[1][5])
freq_step = float(lines[1][6])
num_bins = float(lines[1][8])

print('\nfreq, freq_step, num_bins -->', starting_freq, freq_step, num_bins)

print('\nbeginning bin averaging...')

average_powers = []

for bin_num in range(9, 9+int(num_bins)):
	
	running_average = 0
	for sample_num in range(1, num_samples):
		running_average = running_average + float(lines[sample_num][bin_num])

	average_powers.append(running_average/(num_samples-1))

#print('\n', average_powers)
corrected_powers = []

print('\ncorrecting powers to temp...')
for val in average_powers:
	corrected_powers.append((slope_correction*val) + intercept)

vlsr_av = 0
for i in range(1, num_samples):
	#print(lines[i][int(num_bins)+10])
	vlsr_av = vlsr_av + float(lines[i][int(num_bins)+10])

vlsr_av = vlsr_av/(num_samples-1)
print('\naverage VLSR: ', vlsr_av)


corrected_speeds = []

for i in range(int(num_bins)):
	freq = starting_freq+(i*freq_step)
	speed = -1*vlsr_av - 300000*((freq - 1420.4)/1420.4)
	corrected_speeds.append(speed)


print('\n', corrected_powers)

print('\n', corrected_speeds)


wb = Workbook()
# grab the active worksheet
ws = wb.active
for i in range(len(corrected_powers)):
	ws.cell(column=1, row=i+2, value=corrected_powers[i])

for i in range(len(corrected_speeds)):
	ws.cell(column=2, row=i+2, value=corrected_speeds[i])

wb.save(file_path+'.xslx')
